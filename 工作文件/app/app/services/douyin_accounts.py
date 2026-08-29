from __future__ import annotations

import csv
import base64
import hashlib
import io
import json
import math
import os
import re
import secrets
import sqlite3
import threading
import time
import unicodedata
import zipfile
from contextlib import closing
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Literal

from pydantic import BaseModel, ConfigDict, Field, field_validator, model_validator
from openpyxl import load_workbook


ConnectionMethod = Literal["creator_export"]
_SCHEMA_LOCK = threading.RLock()
_MAX_ROWS = 5000


def _now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def _normalise_header(value: str) -> str:
    text = unicodedata.normalize("NFKC", value or "").strip().lower()
    return re.sub(r"[\s_\-/（）()]+", "", text)


_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "date": ("日期", "时间", "统计日期", "数据日期", "发布时间", "发表时间", "投稿时间", "date"),
    "posts": ("投稿量", "发布作品数", "作品数", "发布数"),
    "views": ("总播放量", "播放量", "播放次数"),
    "likes": ("总点赞量", "点赞量", "点赞数"),
    "comments": ("总评论量", "评论量", "评论数"),
    "shares": ("总分享量", "分享量", "分享数"),
    "five_second_completion": ("5秒完播率", "五秒完播率", "5s完播率"),
    "two_second_bounce": ("2秒跳出率", "两秒跳出率", "2s跳出率"),
    "cover_click_rate": ("封面点击率", "封面点击转化率"),
    "avg_watch_seconds": ("平均播放时长", "平均观看时长"),
    "followers_total": ("总粉丝量", "粉丝总量", "粉丝数"),
    "follower_net": ("粉丝净增", "净增粉丝", "粉丝净增长"),
    "follower_gain": ("新增粉丝", "涨粉量", "新增粉丝数"),
    "follower_loss": ("取关量", "掉粉量", "取关数"),
    "returning_followers": ("回访粉丝量", "回访粉丝数"),
}
_ALIAS_LOOKUP = {
    _normalise_header(alias): key
    for key, aliases in _HEADER_ALIASES.items()
    for alias in aliases
}
_COUNT_KEYS = {
    "posts",
    "views",
    "likes",
    "comments",
    "shares",
    "follower_net",
    "follower_gain",
    "follower_loss",
    "returning_followers",
}
_AVERAGE_KEYS = {
    "five_second_completion",
    "two_second_bounce",
    "cover_click_rate",
    "avg_watch_seconds",
    "followers_total",
}
_RATE_KEYS = {
    "five_second_completion",
    "two_second_bounce",
    "cover_click_rate",
}


class DouyinAccountCreate(BaseModel):
    model_config = ConfigDict(extra="forbid", str_strip_whitespace=True)

    display_name: str = Field(min_length=1, max_length=200)
    douyin_id: str | None = Field(default=None, max_length=200)
    strategy_notes: str = Field(default="", max_length=20_000)
    connection_method: ConnectionMethod = "creator_export"

    @field_validator("douyin_id")
    @classmethod
    def _empty_id_is_none(cls, value: str | None) -> str | None:
        return value or None


class DouyinAccountUpdate(BaseModel):
    model_config = ConfigDict(extra="forbid", str_strip_whitespace=True)

    display_name: str | None = Field(default=None, min_length=1, max_length=200)
    douyin_id: str | None = Field(default=None, max_length=200)
    strategy_notes: str | None = Field(default=None, max_length=20_000)

    @model_validator(mode="after")
    def _at_least_one_field(self) -> DouyinAccountUpdate:
        if not self.model_fields_set:
            raise ValueError("至少提供一个需要更新的账号字段")
        return self


class CreatorDataImport(BaseModel):
    model_config = ConfigDict(extra="forbid", str_strip_whitespace=True)

    filename: str = Field(min_length=1, max_length=260)
    csv_text: str | None = Field(default=None, min_length=1, max_length=2_000_000)
    file_base64: str | None = Field(default=None, min_length=1, max_length=8_500_000)

    @model_validator(mode="after")
    def _one_file_payload(self) -> CreatorDataImport:
        if bool(self.csv_text) == bool(self.file_base64):
            raise ValueError("csv_text 与 file_base64 必须且只能提供一个")
        return self


class DouyinAccountNotFoundError(Exception):
    def __init__(self, account_id: str) -> None:
        super().__init__(f"未找到抖音账号档案：{account_id}")


class DouyinAccountValidationError(Exception):
    """创作者中心导出文件没有可用的指标数据。"""


class DouyinAccountStorageError(Exception):
    """账号档案或指标存储不可用。"""


def _parse_number(value: object) -> float | None:
    text = unicodedata.normalize("NFKC", str(value or "")).strip()
    if not text or text in {"-", "--", "—", "暂无", "null", "None"}:
        return None
    multiplier = 1.0
    if text.endswith("万"):
        multiplier = 10_000.0
        text = text[:-1]
    elif text.endswith("亿"):
        multiplier = 100_000_000.0
        text = text[:-1]
    text = text.rstrip("%秒人次").replace(",", "").replace(" ", "")
    try:
        number = float(text) * multiplier
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def _parse_metric(key: str, value: object) -> float | None:
    number = _parse_number(value)
    if number is None:
        return None
    raw_text = str(value or "")
    if key in _RATE_KEYS and "%" not in raw_text and 0 <= number <= 1:
        return number * 100
    return number


def _detect_dialect(text: str) -> csv.Dialect:
    sample = text[:8192]
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        return csv.excel


def _without_summary_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if len(rows) < 2:
        return rows
    filtered = [
        row
        for row in rows
        if not re.search(r"总计|合计|汇总|全部", str(row.get("date") or ""))
    ]
    return filtered or rows


def _ordered_rows(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], str]:
    dated: list[tuple[datetime, dict[str, Any]]] = []
    for row in rows:
        text = unicodedata.normalize("NFKC", str(row.get("date") or "")).strip()
        if not text:
            return rows, "file_order"
        candidate = text.replace("年", "-").replace("月", "-").replace("日", "")
        parsed: datetime | None = None
        try:
            parsed = datetime.fromisoformat(candidate)
        except ValueError:
            for pattern in ("%Y/%m/%d", "%Y.%m.%d"):
                try:
                    parsed = datetime.strptime(candidate, pattern)
                    break
                except ValueError:
                    continue
        if parsed is None:
            return rows, "file_order"
        dated.append((parsed, row))
    return [row for _, row in sorted(dated, key=lambda item: item[0])], "date_ascending"


def parse_creator_export(csv_text: str) -> dict[str, Any]:
    reader = csv.DictReader(io.StringIO(csv_text.lstrip("\ufeff")), dialect=_detect_dialect(csv_text))
    raw_headers = [str(item or "").strip() for item in (reader.fieldnames or [])]
    header_map = {
        header: _ALIAS_LOOKUP.get(_normalise_header(header))
        for header in raw_headers
    }
    recognised = {header: key for header, key in header_map.items() if key}
    metric_headers = {header: key for header, key in recognised.items() if key != "date"}
    if not metric_headers:
        raise DouyinAccountValidationError(
            "文件中未识别到创作者中心指标列；请从数据中心导出 CSV 后重试。"
        )

    rows: list[dict[str, Any]] = []
    for index, raw in enumerate(reader):
        if index >= _MAX_ROWS:
            raise DouyinAccountValidationError(f"导出文件超过 {_MAX_ROWS} 行上限。")
        item: dict[str, Any] = {"row": index + 1}
        date_header = next((header for header, key in recognised.items() if key == "date"), None)
        if date_header and str(raw.get(date_header) or "").strip():
            item["date"] = str(raw.get(date_header) or "").strip()
        for header, key in metric_headers.items():
            number = _parse_metric(str(key), raw.get(header))
            if number is not None:
                item[str(key)] = number
        if len(item) > (2 if "date" in item else 1):
            rows.append(item)

    rows = _without_summary_rows(rows)
    if not rows:
        raise DouyinAccountValidationError("文件包含指标列，但没有可解析的数值。")
    return {
        "rows": rows,
        "source_row_count": len(rows),
        "recognised_headers": recognised,
        "unrecognised_headers": [header for header in raw_headers if header not in recognised],
    }


def parse_creator_workbook(file_base64: str) -> dict[str, Any]:
    try:
        raw = base64.b64decode(file_base64, validate=True)
    except (ValueError, TypeError) as exc:
        raise DouyinAccountValidationError("XLSX 文件编码无效。") from exc
    if len(raw) > 6_000_000:
        raise DouyinAccountValidationError("XLSX 文件超过 6 MB 上限。")
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            entries = archive.infolist()
            expanded_bytes = sum(item.file_size for item in entries)
            if len(entries) > 1000 or expanded_bytes > 50_000_000:
                raise DouyinAccountValidationError("XLSX 解压后超过安全上限。")
    except zipfile.BadZipFile as exc:
        raise DouyinAccountValidationError("XLSX 文件不是有效的工作簿。") from exc
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise DouyinAccountValidationError("XLSX 文件无法打开或结构无效。") from exc

    try:
        for sheet in workbook.worksheets:
            all_rows = sheet.iter_rows(values_only=True)
            header_values: list[str] | None = None
            header_map: dict[int, str] = {}
            data_rows: list[dict[str, Any]] = []
            for row_index, values in enumerate(all_rows, start=1):
                cells = [str(value).strip() if value is not None else "" for value in values]
                if header_values is None:
                    candidate = {
                        index: key
                        for index, header in enumerate(cells)
                        if (key := _ALIAS_LOOKUP.get(_normalise_header(header)))
                    }
                    if any(key != "date" for key in candidate.values()):
                        header_values = cells
                        header_map = candidate
                    elif row_index >= 30:
                        break
                    continue
                if len(data_rows) >= _MAX_ROWS:
                    raise DouyinAccountValidationError(f"导出文件超过 {_MAX_ROWS} 行上限。")
                item: dict[str, Any] = {"row": row_index}
                for index, key in header_map.items():
                    value = values[index] if index < len(values) else None
                    if key == "date":
                        if value is not None and str(value).strip():
                            item["date"] = str(value).strip()
                    else:
                        number = _parse_metric(key, value)
                        if number is not None:
                            item[key] = number
                if len(item) > (2 if "date" in item else 1):
                    data_rows.append(item)
            data_rows = _without_summary_rows(data_rows)
            if header_values is not None and data_rows:
                recognised = {
                    header_values[index]: key
                    for index, key in header_map.items()
                    if header_values[index]
                }
                return {
                    "rows": data_rows,
                    "source_row_count": len(data_rows),
                    "recognised_headers": recognised,
                    "unrecognised_headers": [
                        header
                        for index, header in enumerate(header_values)
                        if header and index not in header_map
                    ],
                    "worksheet": sheet.title,
                }
    finally:
        workbook.close()
    raise DouyinAccountValidationError(
        "工作簿中未识别到创作者中心指标表；请确认导出的是作品或粉丝数据。"
    )


def _aggregate(rows: list[dict[str, Any]]) -> dict[str, float]:
    result: dict[str, float] = {}
    for key in _COUNT_KEYS:
        values = [float(row[key]) for row in rows if isinstance(row.get(key), (int, float))]
        if values:
            result[key] = round(sum(values), 3)
    for key in _AVERAGE_KEYS:
        values = [float(row[key]) for row in rows if isinstance(row.get(key), (int, float))]
        if values:
            result[key] = round(sum(values) / len(values), 3)
    views = result.get("views")
    posts = result.get("posts")
    interactions = sum(result.get(key, 0.0) for key in ("likes", "comments", "shares"))
    if views and views > 0:
        result["interaction_rate"] = round(interactions / views * 100, 3)
    if posts and posts > 0:
        if views is not None:
            result["views_per_post"] = round(views / posts, 3)
        if "follower_net" in result:
            result["follower_net_per_post"] = round(result["follower_net"] / posts, 3)
    return result


def _comparison(previous: dict[str, float], current: dict[str, float], key: str) -> dict[str, float] | None:
    if key not in previous or key not in current:
        return None
    before = previous[key]
    after = current[key]
    return {
        "previous": before,
        "current": after,
        "delta": round(after - before, 3),
        "relative_change_percent": round((after - before) / abs(before) * 100, 3) if before else 0.0,
    }


def analyse_creator_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    if not rows:
        return {"status": "insufficient", "row_count": 0, "recommendations": []}
    ordered_rows, order_basis = _ordered_rows(rows)
    split = max(1, len(ordered_rows) // 2)
    previous_rows = ordered_rows[:split] if len(ordered_rows) > 1 else []
    current_rows = ordered_rows[split:] if len(ordered_rows) > 1 else ordered_rows
    previous = _aggregate(previous_rows)
    current = _aggregate(current_rows)
    comparisons = {
        key: comparison
        for key in (
            "views_per_post",
            "interaction_rate",
            "five_second_completion",
            "two_second_bounce",
            "cover_click_rate",
            "avg_watch_seconds",
            "follower_net_per_post",
        )
        if (comparison := _comparison(previous, current, key)) is not None
    }

    recommendations: list[dict[str, str]] = []

    def add(area: str, finding: str, action: str, evidence: str) -> None:
        recommendations.append(
            {"area": area, "finding": finding, "action": action, "evidence": evidence}
        )

    bounce = comparisons.get("two_second_bounce")
    if bounce and bounce["delta"] >= 3:
        add(
            "开头 2 秒",
            "最近一段的 2 秒跳出率上升。",
            "下一条只改首帧和第一句话，直接展示结果或冲突点，其他变量保持不变。",
            f"较前一段上升 {bounce['delta']:.1f} 个百分点",
        )
    completion = comparisons.get("five_second_completion")
    if completion and completion["delta"] <= -3:
        add(
            "前 5 秒承接",
            "最近一段的 5 秒完播率下降。",
            "删掉开场铺垫，在第 5 秒前交代观众能得到什么，并登记一次单变量实验。",
            f"较前一段下降 {abs(completion['delta']):.1f} 个百分点",
        )
    cover = comparisons.get("cover_click_rate")
    if cover and cover["delta"] <= -2:
        add(
            "封面与标题",
            "最近一段的封面点击率下降。",
            "用同一内容测试两版封面文案，只改变利益点表达，不同时改脚本。",
            f"较前一段下降 {abs(cover['delta']):.1f} 个百分点",
        )
    watch = comparisons.get("avg_watch_seconds")
    if watch and watch["relative_change_percent"] <= -10:
        add(
            "中段节奏",
            "最近一段的平均播放时长明显下降。",
            "把核心证明提前，删除重复解释，并对照完整口播逐段压缩。",
            f"较前一段下降 {abs(watch['relative_change_percent']):.1f}%",
        )
    interaction = comparisons.get("interaction_rate")
    if interaction and interaction["relative_change_percent"] <= -10:
        add(
            "互动设计",
            "最近一段每次播放带来的互动下降。",
            "结尾改为一个具体、容易回答的问题，并预先准备两条真实评论回复。",
            f"互动率较前一段下降 {abs(interaction['relative_change_percent']):.1f}%",
        )
    views = comparisons.get("views_per_post")
    if views and views["relative_change_percent"] <= -15:
        add(
            "选题稳定性",
            "最近一段单条平均播放量下降。",
            "回看同期最高播放作品，复用问题类型和证据结构，不复用原句。",
            f"单条平均播放较前一段下降 {abs(views['relative_change_percent']):.1f}%",
        )
    followers = comparisons.get("follower_net_per_post")
    if followers and followers["relative_change_percent"] <= -20:
        add(
            "关注转化",
            "最近一段单条净增粉下降。",
            "明确账号持续提供的固定价值，并让结尾承接下一期内容。",
            f"单条净增粉较前一段下降 {abs(followers['relative_change_percent']):.1f}%",
        )

    if len(ordered_rows) < 2:
        add(
            "数据积累",
            "当前只有一个观测点，无法判断趋势。",
            "保持相同导出口径，至少再导入一个相邻周期后再比较。",
            "趋势样本不足",
        )
    elif not recommendations:
        add(
            "下一轮实验",
            "当前可比较指标没有出现明显下滑。",
            "保持表现稳定的结构，只选择一个变量做下一轮实验并回填同一观察窗。",
            "仅基于账号自身前后两段数据，不使用外部行业基准",
        )

    return {
        "status": "ready" if len(ordered_rows) >= 2 else "insufficient",
        "row_count": len(ordered_rows),
        "order_basis": order_basis,
        "previous_row_count": len(previous_rows),
        "current_row_count": len(current_rows),
        "previous": previous,
        "current": current,
        "comparisons": comparisons,
        "recommendations": recommendations,
        "evidence_boundary": "只比较该账号自身相邻两段数据，不代表行业基准或因果结论。",
    }


def _analysis_response_state(analysis: object) -> tuple[str, str]:
    analysis_status = (
        str(analysis.get("status") or "")
        if isinstance(analysis, dict)
        else ""
    )
    if analysis_status in {"ready", "completed"}:
        return "completed", "已按账号自身相邻时段完成趋势比较。"
    return (
        "insufficient",
        "当前缺少可比较的相邻时段，尚不能判断趋势；请至少提供两个可比较观测点。",
    )


class DouyinAccountStore:
    """SQLite/WAL 账号档案和创作者中心导出数据。"""

    def __init__(self) -> None:
        self._root: Path | None = None
        self._initialised = False

    @property
    def root(self) -> Path:
        if self._root is None:
            configured = os.environ.get("PROJECT024_DOUYIN_OPERATIONS_ROOT", "").strip()
            self._root = (
                Path(configured).expanduser().resolve()
                if configured
                else Path(__file__).resolve().parents[2] / "var" / "douyin_operations"
            )
            self._root.mkdir(parents=True, exist_ok=True)
        return self._root

    @property
    def database_path(self) -> Path:
        return self.root / "douyin_operations.sqlite3"

    def _open_connection(self) -> sqlite3.Connection:
        connection = sqlite3.connect(self.database_path, timeout=30.0, isolation_level=None)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA busy_timeout = 30000")
        return connection

    def _ensure_initialised(self) -> None:
        if self._initialised:
            return
        with _SCHEMA_LOCK:
            if self._initialised:
                return
            for attempt in range(4):
                try:
                    with closing(self._open_connection()) as connection:
                        connection.execute("PRAGMA journal_mode = WAL")
                        connection.execute("PRAGMA synchronous = FULL")
                        connection.execute(
                            """
                            CREATE TABLE IF NOT EXISTS douyin_accounts (
                                id TEXT PRIMARY KEY,
                                douyin_id TEXT UNIQUE,
                                created_at TEXT NOT NULL,
                                updated_at TEXT NOT NULL,
                                version INTEGER NOT NULL,
                                record_json TEXT NOT NULL
                            )
                            """
                        )
                        connection.execute(
                            """
                            CREATE TABLE IF NOT EXISTS douyin_metric_imports (
                                id TEXT PRIMARY KEY,
                                account_id TEXT NOT NULL,
                                fingerprint TEXT NOT NULL,
                                imported_at TEXT NOT NULL,
                                record_json TEXT NOT NULL,
                                UNIQUE (account_id, fingerprint),
                                FOREIGN KEY (account_id) REFERENCES douyin_accounts(id)
                            )
                            """
                        )
                    self._initialised = True
                    return
                except sqlite3.OperationalError as exc:
                    if "locked" not in str(exc).lower() or attempt == 3:
                        raise
                    time.sleep(0.05 * (attempt + 1))

    @staticmethod
    def _load(row: sqlite3.Row) -> dict[str, Any]:
        try:
            record = json.loads(str(row["record_json"]))
        except (TypeError, ValueError) as exc:
            raise DouyinAccountStorageError("抖音账号数据无法解析") from exc
        if not isinstance(record, dict):
            raise DouyinAccountStorageError("抖音账号数据结构无效")
        return record

    @staticmethod
    def _dump(record: dict[str, Any]) -> str:
        return json.dumps(record, ensure_ascii=False, sort_keys=True, separators=(",", ":"), allow_nan=False)

    @staticmethod
    def _next_id(prefix: str) -> str:
        return f"{prefix}_{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}_{secrets.token_hex(4)}"

    def create(self, payload: DouyinAccountCreate) -> dict[str, Any]:
        self._ensure_initialised()
        now = _now()
        record: dict[str, Any] = {
            "schema_version": 1,
            "id": self._next_id("dya"),
            "platform": "douyin",
            "display_name": payload.display_name,
            "douyin_id": payload.douyin_id,
            "strategy_notes": payload.strategy_notes,
            "connection_method": payload.connection_method,
            "official_authorized": False,
            "created_at": now,
            "updated_at": now,
            "version": 1,
        }
        with closing(self._open_connection()) as connection:
            try:
                connection.execute("BEGIN IMMEDIATE")
                connection.execute(
                    "INSERT INTO douyin_accounts (id, douyin_id, created_at, updated_at, version, record_json) VALUES (?, ?, ?, ?, ?, ?)",
                    (record["id"], record["douyin_id"], now, now, 1, self._dump(record)),
                )
                connection.execute("COMMIT")
                record["deduplicated"] = False
                return record
            except sqlite3.IntegrityError:
                connection.execute("ROLLBACK")
                if payload.douyin_id:
                    row = connection.execute(
                        "SELECT record_json FROM douyin_accounts WHERE douyin_id = ?",
                        (payload.douyin_id,),
                    ).fetchone()
                    if row is not None:
                        existing = self._load(row)
                        existing["deduplicated"] = True
                        return existing
                raise DouyinAccountStorageError("抖音账号档案写入冲突")
            except sqlite3.Error as exc:
                if connection.in_transaction:
                    connection.execute("ROLLBACK")
                raise DouyinAccountStorageError("抖音账号档案暂时无法保存") from exc

    def list_all(self) -> list[dict[str, Any]]:
        self._ensure_initialised()
        try:
            with closing(self._open_connection()) as connection:
                rows = connection.execute(
                    "SELECT record_json FROM douyin_accounts ORDER BY updated_at DESC, id DESC"
                ).fetchall()
        except sqlite3.Error as exc:
            raise DouyinAccountStorageError("抖音账号档案暂时无法读取") from exc
        return [self._load(row) for row in rows]

    def get(self, account_id: str) -> dict[str, Any]:
        self._ensure_initialised()
        with closing(self._open_connection()) as connection:
            row = connection.execute(
                "SELECT record_json FROM douyin_accounts WHERE id = ?", (account_id,)
            ).fetchone()
        if row is None:
            raise DouyinAccountNotFoundError(account_id)
        return self._load(row)

    def update(self, account_id: str, payload: DouyinAccountUpdate) -> dict[str, Any]:
        self._ensure_initialised()
        with closing(self._open_connection()) as connection:
            try:
                connection.execute("BEGIN IMMEDIATE")
                row = connection.execute(
                    "SELECT version, record_json FROM douyin_accounts WHERE id = ?", (account_id,)
                ).fetchone()
                if row is None:
                    connection.execute("ROLLBACK")
                    raise DouyinAccountNotFoundError(account_id)
                record = self._load(row)
                changes = payload.model_dump(exclude_unset=True)
                record.update(changes)
                record["updated_at"] = _now()
                record["version"] = int(row["version"]) + 1
                connection.execute(
                    "UPDATE douyin_accounts SET douyin_id = ?, updated_at = ?, version = ?, record_json = ? WHERE id = ?",
                    (
                        record.get("douyin_id"),
                        record["updated_at"],
                        record["version"],
                        self._dump(record),
                        account_id,
                    ),
                )
                connection.execute("COMMIT")
                return record
            except DouyinAccountNotFoundError:
                raise
            except sqlite3.IntegrityError as exc:
                if connection.in_transaction:
                    connection.execute("ROLLBACK")
                raise DouyinAccountStorageError("该抖音号已绑定其他本地档案") from exc
            except sqlite3.Error as exc:
                if connection.in_transaction:
                    connection.execute("ROLLBACK")
                raise DouyinAccountStorageError("抖音账号档案暂时无法更新") from exc

    def import_creator_data(self, account_id: str, payload: CreatorDataImport) -> dict[str, Any]:
        self.get(account_id)
        parsed = (
            parse_creator_export(payload.csv_text)
            if payload.csv_text is not None
            else parse_creator_workbook(str(payload.file_base64))
        )
        canonical = json.dumps(
            {"rows": parsed["rows"], "headers": parsed["recognised_headers"]},
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
            allow_nan=False,
        )
        fingerprint = hashlib.sha256(canonical.encode("utf-8")).hexdigest()
        imported_at = _now()
        analysis = analyse_creator_rows(parsed["rows"])
        status, message = _analysis_response_state(analysis)
        record: dict[str, Any] = {
            "schema_version": 1,
            "id": self._next_id("dyi"),
            "account_id": account_id,
            "filename": Path(payload.filename).name,
            "fingerprint": fingerprint,
            "imported_at": imported_at,
            **parsed,
            "status": status,
            "message": message,
            "analysis": analysis,
        }
        with closing(self._open_connection()) as connection:
            try:
                connection.execute("BEGIN IMMEDIATE")
                connection.execute(
                    "INSERT INTO douyin_metric_imports (id, account_id, fingerprint, imported_at, record_json) VALUES (?, ?, ?, ?, ?)",
                    (record["id"], account_id, fingerprint, imported_at, self._dump(record)),
                )
                connection.execute(
                    "UPDATE douyin_accounts SET updated_at = ? WHERE id = ?",
                    (imported_at, account_id),
                )
                connection.execute("COMMIT")
                record["deduplicated"] = False
                return record
            except sqlite3.IntegrityError:
                connection.execute("ROLLBACK")
                row = connection.execute(
                    "SELECT record_json FROM douyin_metric_imports WHERE account_id = ? AND fingerprint = ?",
                    (account_id, fingerprint),
                ).fetchone()
                if row is None:
                    raise DouyinAccountStorageError("创作者中心数据写入冲突")
                existing = self._load(row)
                status, message = _analysis_response_state(existing.get("analysis"))
                existing["status"] = status
                existing["message"] = message
                existing["deduplicated"] = True
                return existing
            except sqlite3.Error as exc:
                if connection.in_transaction:
                    connection.execute("ROLLBACK")
                raise DouyinAccountStorageError("创作者中心数据暂时无法保存") from exc

    def list_imports(self, account_id: str) -> list[dict[str, Any]]:
        self.get(account_id)
        try:
            with closing(self._open_connection()) as connection:
                rows = connection.execute(
                    "SELECT record_json FROM douyin_metric_imports WHERE account_id = ? ORDER BY imported_at DESC, id DESC",
                    (account_id,),
                ).fetchall()
        except sqlite3.Error as exc:
            raise DouyinAccountStorageError("创作者中心数据暂时无法读取") from exc
        return [self._load(row) for row in rows]

    def analysis(self, account_id: str) -> dict[str, Any]:
        account = self.get(account_id)
        imports = self.list_imports(account_id)
        if not imports:
            return {
                "account": account,
                "status": "needs_data",
                "message": "尚未导入创作者中心数据。",
                "latest_import": None,
                "analysis": None,
            }
        latest = imports[0]
        analysis = latest.get("analysis")
        status, message = _analysis_response_state(analysis)
        return {
            "account": account,
            "status": status,
            "message": message,
            "latest_import": {
                key: latest.get(key)
                for key in (
                    "id",
                    "filename",
                    "imported_at",
                    "source_row_count",
                    "recognised_headers",
                    "deduplicated",
                )
                if latest.get(key) is not None
            },
            "analysis": analysis,
        }
